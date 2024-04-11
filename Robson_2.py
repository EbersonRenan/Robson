#https://www.youtube.com/watch?v=gjaiAYMhi3U&t=76s
#playwright codegen https://www.sicredi.com.br/home/
from playwright.sync_api import Playwright, sync_playwright
from datetime import datetime
import calendar
import openpyxl
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook
import os
import shutil
from pathlib import Path
import sys

#Carregue a pasta de trabalho do Excel.
workbook = load_workbook('C:\python\CONTABILIDADE\Pasta1.xlsx')  #Tem que estar na mesma pasta


planilha = workbook['Planilha1']
planilha = workbook.active

# Abre a aba.
sheet = workbook.get_sheet_by_name('Planilha1')

# Obtenha os valores da coluna A.
column_a_values = [cell.value for cell in sheet['A']]
column_k_values = [cell.value for cell in sheet['L']]

# Inicializar a variável caminho com uma string vazia
def Abrir_Excel():
    excel = 'C:\python\CONTABILIDADE\Pasta1.xlsx'
    os.startfile(excel)

def buscar_caminho():
    """Busca o caminho da pasta para salvar um arquivo em Python no Tkinter.

    Returns:
    O caminho da pasta selecionada pelo usuário.
    """
    root = tk.Tk()
    root.withdraw()
    
    diretorio = filedialog.askdirectory(title="Selecione a pasta para salvar o arquivo")
    
    global caminho
    caminho = diretorio  
    #print(caminho)

def limpar():
    # Limpar o widget NomeUsuario_input
    NomeUsuario_input.delete(0, tk.END)

    # Limpar o widget start_month_input
    start_month_input.delete(0, tk.END)

    # Limpar o widget end_month_input
    end_month_input.delete(0, tk.END)

    # Limpar o widget year_input
    year_input.delete(0, tk.END)

    # Limpar o widget bank_input
    bank_input.delete(0, tk.END)

    # Limpar o widget empresas_input
    empresas_input.delete(0, tk.END)


def salvar():

    # Pega o input do usuário
    NomeUsuario = NomeUsuario_input.get()
    mes_inicial = start_month_input.get()
    mes_final = end_month_input.get()
    ano = year_input.get()
    banco = bank_input.get()
    empresa = empresas_input.get()
    salvar = caminho
    MostrarExecução = var.get()
    # Cria um arquivo de texto
    with open("C:\python\CONTABILIDADE\input.txt", "w") as f:
             
        # Escreve o input do usuário no arquivo
        f.write("Nome do usuario: {}\n".format(NomeUsuario))
        f.write("Mês inicial: {}\n".format(mes_inicial))
        f.write("Mês final: {}\n".format(mes_final))
        f.write("Ano: {}\n".format(ano))
        f.write("Banco: {}\n".format(banco))
        f.write("Empresa: {}\n".format(empresa))
        f.write("Salvar: {}\n".format(salvar.replace(":",";")))
        f.write("Mostrar Execução: {}\n".format(MostrarExecução))
        
    window = tk.Tk()
    window.title("Robson")
    window.geometry("220x100")
    def on_click():
        window.destroy()

# Defina a cor de fundo para laranja
    window.configure(bg="orange")
    CONCLUIDO_label = tk.Label(window, text="Concluido", bg="orange", font=("Arial", 30, "bold"), fg="white")
    CONCLUIDO_label.grid(row=1, column=0, columnspan=3)

    button = tk.Button(window, text="OK", command=on_click, font=("Arial", 16, "bold"))
    button.grid(row=2, column=1)


def Gerados():

    window = tk.Tk()
    window.title("Robson")
    window.geometry("220x100")
    def on_click():
        window.destroy()
        sys.exit()
# Defina a cor de fundo para laranja
    window.configure(bg="orange")
    CONCLUIDO_label = tk.Label(window, text="Concluido", bg="orange", font=("Arial", 30, "bold"), fg="white")
    CONCLUIDO_label.grid(row=1, column=0, columnspan=3)

    button = tk.Button(window, text="OK", command=on_click, font=("Arial", 16, "bold"))
    button.grid(row=2, column=1)

def Execucao_Robson():
   
        #Ler o txt com os conmandos para execução do robo
        with open("C:\python\CONTABILIDADE\input.txt", "r") as f:
        # Obter uma lista de strings com cada linha do arquivo
            linhas = f.read().splitlines()
        # Criar um dicionário vazio para armazenar as variáveis
        variaveis = {}
        # Percorrer a lista de linhas
        for linha in linhas:
            # Separar o nome e o valor da variável usando o caractere ":"
            nome, valor = linha.split(":")
            # Remover os espaços em branco do nome e do valor
            nome = nome.strip()
            nome = nome.replace("ê", "e")
            nome = nome.replace(" ", "_")
            valor = valor.strip()
            # Converter o valor da variável em seu tipo correspondente usando a função eval
            #valor = eval(valor)
            # Adicionar a variável ao dicionário usando o nome como chave e o valor como valor
            variaveis[nome] = valor
        # Imprimir o dicionário de variáveis na tela
        #print(variaveis)

        MesInicial = variaveis["Mes_inicial"]
        MesFinal = variaveis["Mes_final"]
        Ano = variaveis["Ano"]
        Banco = variaveis["Banco"]
        Empresa = variaveis["Empresa"]
        Salvar = variaveis["Salvar"]
        Salvar = Salvar.replace(";",":")
        MostrarExecução = variaveis["Mostrar_Execução"]

        if MostrarExecução == "1":
            headless = False
        else:
            headless = True



        lista_de_mes = list(range(int(MesInicial), int(MesFinal)+1))
        l = 0

        MesLimite = datetime.now().month  #verifica o mes

        if MesLimite == 1:
            MesLimite = 12
        else:
            MesLimite = MesLimite - 1
        #Acresenta um 0 na frente de numeros menores que 10
        if MesLimite < 10:
            MesLimitel = "0" + str(MesLimite)
        else:
            MesLimite = int(MesLimite)

        #VERIFICA AS VARIAVEIS NA PLANILHA
        for row in range(2, planilha.max_row + 1):
    
            #Coluna nome da empresa
            celula = planilha.cell(row=row, column=1)
            empresa = celula.value

            #COLUNA CNPJ/CPF
            celula = planilha.cell(row=row, column=2)
            CNPJ_CPF = celula.value

            #Coluna login
            celula = planilha.cell(row=row, column=3)
            Login = celula.value

            #coluna senha
            celula = planilha.cell(row=row, column=4)
            SenhaB = celula.value

            #coluna banco
            celula = planilha.cell(row=row, column=5)
            BancoPlanilha = celula.value
            
            #if Banco == 'Todos':
                #print(empresa, CNPJ_CPF, Login, SenhaB, BancoPlanilha)
            #elif BancoPlanilha == Banco:
                #print(empresa, CNPJ_CPF , Login, SenhaB)

            def Banrisul(playwright: Playwright) -> None:   ##BANRISUL
                    print()
                    print("Banrisul")
                    print()
                    workbook = openpyxl.load_workbook("C:\python\CONTABILIDADE\Pasta1.xlsx")

                    planilha = workbook['Planilha1']
                    planilha = workbook.active
                        
                    for row in range(2, planilha.max_row + 1):
                
                    #Coluna nome da empresa
                        celula = planilha.cell(row=row, column=1)
                        empresa = celula.value
                        #empresa3 = empresa
                       

                        #COLUNA CNPJ/CPF
                        celula = planilha.cell(row=row, column=2)
                        CNPJ_CPF = celula.value

                        #Coluna login
                        celula = planilha.cell(row=row, column=3)
                        Login = celula.value

                        #coluna senha
                        celula = planilha.cell(row=row, column=4)
                        SenhaB = celula.value
                        SenhaB = str(SenhaB)

                        #coluna banco
                        celula = planilha.cell(row=row, column=5)
                        BancoPlanilha = celula.value

                        #coluna observação
                        celula = planilha.cell(row=row, column=6)
                        Obs = celula.value

                        print(empresa)
                        Empresa2 = Empresa.replace(' ', '')
                        
                        if empresa.replace(' ', '') == Empresa2:
                            if BancoPlanilha == 'Banrisul':
                                print(empresa)
                                CNPJ_CPF = CNPJ_CPF.replace('.','')
                                CNPJ_CPF = CNPJ_CPF.replace('/','')
                                CNPJ_CPF = CNPJ_CPF.replace('-','')
                                lista_de_mes = list(range(int(MesInicial), int(MesFinal)+1))
                                l = 0
                                for l in lista_de_mes:
                                    print(l)
                                    if l < 10:
                                        DataL = "0" + str(l) +"/"+ Ano
                                        l = "0" + str(l)
                                    else:
                                        DataL = str(l) +"/"+ Ano
                                    print(DataL)

                                    AnoDoMes = datetime.now().year

                                    #if Ano != AnoDoMes:
                                    #    AnoDoMes = Ano 

                                    limitador = [] 
                                    if MesLimite < 10:
                                        limitador = "0" + str(MesLimite) + "/" + str(Ano)
                                    else:
                                        limitador = str(MesLimite) + "/" + str(Ano)
                                    
                                    #print(limitador)
                                    res = calendar.monthrange(int(Ano), int(l))
                                    res[0]
                                    res[1]
                                    ultimoDia = res[1]
                                    DataInicial = "01" + "/" + str(l) + "/" + str(Ano)
                                    DataFinal = str(ultimoDia) + "/" + str(l) + "/" + str(Ano)
                                    MesdeSelecao = str(l) + "/" + str(Ano)

                                    browser = playwright.chromium.launch(channel="chrome", headless=headless)
                                    context = browser.new_context()
                                    page = context.new_page()
                                    #ABRE A PAGINA E MANIPULA A MESMA
                                    page.goto("https://www.banrisul.com.br/")
                                    with page.expect_popup() as page1_info:
                                        page.frame_locator("frame[name=\"Homepage\"]").get_by_role("link", name="Office Banking").click()
                                    page1 = page1_info.value
                                    page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_placeholder("Operador ou CPF").click()

                                    if Obs == "Keli":
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_placeholder("Operador ou CPF").fill(Login)
                                        page1.wait_for_timeout(1000)
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_placeholder("Operador ou CPF").press("Enter")
                                        
                                        Agora = datetime.now()
                                        HoraAgora = Agora.time().strftime("%H:%M:%S")
                                        MeioDia = "12:00:00"
                                        HoraAgora = str(HoraAgora)
                                        MeioDia = str(MeioDia)

                                        print(HoraAgora)
                                        if HoraAgora < MeioDia:

                                            def indentificar_item(item):
                                                    if isinstance(item, str):
                                                        if item.isalpha():
                                                            return 'Letra'
                                                        elif item.isdigit():
                                                            return 'Numero'
                                                        else:
                                                            return 'simbulo'
                                                    elif isinstance(item, (int, float)):
                                                        return 'Numero'
                                                    else:
                                                        return 'outro'
                                            def senhas():
                                                Senha = SenhaB.swapcase()
                                                print(Senha)
                                                for item in Senha:
                                                    if indentificar_item(item) == 'Letra':
                                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_text(str(item), exact=True).click()
                                                    else:
                                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_text(item).click()
                                            senhas()
                                        else:

                                            def indentificar_item(item):
                                                if isinstance(item, str):
                                                    if item.isalpha():
                                                        return 'Letra'
                                                    elif item.isdigit():
                                                        return 'Numero'
                                                    else:
                                                        return 'simbulo'
                                                elif isinstance(item, (int, float)):
                                                    return 'Numero'
                                                else:
                                                    return 'outro'

                                            Senha = SenhaB.swapcase()
                                            for item in Senha:
                                                if indentificar_item(item) == 'Letra':
                                                    page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_text(str(item), exact=True).click()
                                                else:
                                                    page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_text(item).click()

                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("button", name="Confirmar").click()
                                        
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("link", name=empresa).click()


                                    else:
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_placeholder("Operador ou CPF").fill(CNPJ_CPF)

                                    
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_placeholder("Operador ou CPF").press("Enter")
                                        page1.wait_for_timeout(2000)
                                        
                                        Agora = datetime.now()
                                        HoraAgora = Agora.time().strftime("%H:%M:%S")
                                        MeioDia = "12:00:00"
                                        HoraAgora = str(HoraAgora)
                                        MeioDia = str(MeioDia)

                                        print(HoraAgora)
                                        if HoraAgora < MeioDia:

                                            def indentificar_item(item):
                                                    if isinstance(item, str):
                                                        if item.isalpha():
                                                            return 'Letra'
                                                        elif item.isdigit():
                                                            return 'Numero'
                                                        else:
                                                            return 'simbulo'
                                                    elif isinstance(item, (int, float)):
                                                        return 'Numero'
                                                    else:
                                                        return 'outro'
                                            def senhas():
                                                Senha = SenhaB.swapcase()
                                                print(Senha)
                                                for item in Senha:
                                                    if indentificar_item(item) == 'Letra':
                                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_text(str(item), exact=True).click()
                                                    else:
                                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_text(item).click()
                                            senhas()
                                        else:

                                            def indentificar_item(item):
                                                if isinstance(item, str):
                                                    if item.isalpha():
                                                        return 'Letra'
                                                    elif item.isdigit():
                                                        return 'Numero'
                                                    else:
                                                        return 'simbulo'
                                                elif isinstance(item, (int, float)):
                                                    return 'Numero'
                                                else:
                                                    return 'outro'

                                            Senha = SenhaB.swapcase()
                                            for item in Senha:
                                                if indentificar_item(item) == 'Letra':
                                                    page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_text(str(item), exact=True).click()
                                                else:
                                                    page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_text(item).click()

                                    
                                            page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("button", name="Confirmar").click()

                                    #page1.wait_for_timeout(1000000)
                                    #page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("cell", name="FABRICIO PERIN DALL AGNOL").click()
                                    #page1.get_by_role("button", name="» Li a mensagem e estou ciente do seu conteúdo").click()
                                    page1.frame_locator("iframe[name=\"wMenu\"]").get_by_role("link", name="Extratos").click()
                                    page1.frame_locator("iframe[name=\"wMenu\"]").get_by_role("link", name="Conta-Corrente").click()
                                    page1.wait_for_timeout(1000)
                                    page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("link", name="Período").click()
                                    
                                    limitador = str(limitador)
                                    MesdeSelecao = str(MesdeSelecao)
                                    
                                    if str(limitador).replace("/","") == "002024":
                                        limitador = DataL
                                    
                                    if limitador == MesdeSelecao:        
                                        
                                        periodo = "01/" + limitador
                                        periodo = periodo.replace("/"+Ano, "")

                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("link", name="Período").click()
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_label(f"Últimos 2 meses (período com data inicial maior ou igual a {periodo})").check()
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_label("De").click()
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_label("De").fill(DataInicial)
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_label("Até").click()
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_label("Até").fill(DataFinal)
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("button", name="» Consultar").click()
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("button", name="» Gravar").click()      
                                        page1.get_by_label("OFX", exact=True).check()
                                        #executa o download do ofx
                                        with page1.expect_download() as download_info:
                                            page1.get_by_role("button", name="» Gravar").click()
                                        download = download_info.value
                                        #print(f'Salvando: {download.path()}')
                                        diretorio = os.getcwd()
                                        pasta_final = os.path.join(diretorio, f"download/BANRISUL/OFX/{DataL.replace('/','')}.ofx")
                                        pasta_final = pasta_final.replace("\\", '/')
                                        download.save_as(pasta_final)
                                        origem = pasta_final
                                        destino = f"{Salvar.replace(';',':')}/BANRISUL/{empresa.replace(' ', '_')}/OFX/"
                                        caminho = Path(destino)
                                        existe = caminho.exists () # retorna True ou False
                                        #print(existe)
                                        if existe == False:
                                            os.makedirs (destino) 
                                        destino_completo = os.path.join (destino, os.path.basename (origem))
                                        shutil.move (origem, destino_completo)
                                        
                                        print(f'Salvo na pasta: {pasta_final}')

                                        #executa o download do PDF   
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("button", name="» Gravar").click()      
                                        page1.get_by_label("PDF", exact=True).check()
                                        page1.wait_for_timeout(1000)
                                        with page1.expect_download() as download_info:
                                            page1.get_by_role("button", name="» Gravar").click()
                                        download = download_info.value
                                        #print(f'Salvando: {download.path()}')
                                        diretorio = os.getcwd()
                                        pasta_final = os.path.join(diretorio, f"download/BANRISUL/PDF/{DataL.replace('/','')}.PDF")
                                        pasta_final = pasta_final.replace("\\", '/')
                                        download.save_as(pasta_final)
                                        origem = pasta_final
                                        destino = f"{Salvar.replace(';',':')}/BANRISUL/{empresa.replace(' ', '_')}/PDF/"
                                        caminho = Path(destino)
                                        existe = caminho.exists () # retorna True ou False
                                        #print(existe)
                                        if existe == False:
                                            os.makedirs (destino) 
                                        destino_completo = os.path.join (destino, os.path.basename (origem))
                                        shutil.move (origem, destino_completo)
                                        print(f'Salvo na pasta: {pasta_final}')
                                    else:
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_label("Últimos 2 anos (mensal)").check()        
                                        #seleção de mes
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("combobox", name="Mês/Ano").select_option(DataL)
                                        page1.wait_for_timeout(2000)
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("button", name="» Consultar").click()
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("button", name="» Gravar").click()      
                                        page1.get_by_label("OFX", exact=True).check()
                                        page1.wait_for_timeout(1000)
                                    
                                        #executa o download do ofx 
                                        with page1.expect_download() as download_info:
                                            page1.get_by_role("button", name="» Gravar").click()
                                        download = download_info.value
                                        #print(f'Salvando: {download.path()}')
                                        diretorio = os.getcwd()
                                        pasta_final = os.path.join(diretorio, f"download/BANRISUL/OFX/{DataL.replace('/','')}.ofx")
                                        pasta_final = pasta_final.replace("\\", '/')
                                        download.save_as(pasta_final)
                                        origem = pasta_final
                                        destino = f"{Salvar.replace(';',':')}/BANRISUL/{empresa.replace(' ', '_')}/OFX/"
                                        destino.replace(' ', '')
                                        caminho = Path(destino)
                                        existe = caminho.exists () # retorna True ou False
                                        #print(existe)
                                        if existe == False:
                                            os.makedirs (destino) 
                                        destino_completo = os.path.join (destino, os.path.basename (origem))
                                        shutil.move (origem, destino_completo)
                                        print(f'Salvo na pasta: {pasta_final}')

                                        #executa o download do PDF   
                                        page1.frame_locator("iframe[name=\"wProgramas\"]").get_by_role("button", name="» Gravar").click()      
                                        page1.get_by_label("PDF", exact=True).check()
                                        page1.wait_for_timeout(1000)
                                        with page1.expect_download() as download_info:
                                            page1.get_by_role("button", name="» Gravar").click()
                                        download = download_info.value
                                        #print(f'Salvando: {download.path()}')
                                        diretorio = os.getcwd()
                                        pasta_final = os.path.join(diretorio, f"download/BANRISUL/PDF/{DataL.replace('/','')}.PDF")
                                        pasta_final = pasta_final.replace("\\", '/')
                                        download.save_as(pasta_final)
                                        origem = pasta_final
                                        destino = f"{Salvar.replace(';',':')}/BANRISUL/{empresa.replace(' ', '_')}/PDF/"
                                        destino.replace(' ', '')
                                        caminho = Path(destino)
                                        existe = caminho.exists() # retorna True ou False
                                        #print(existe)
                                        if existe == False:
                                            os.makedirs(destino) 
                                        destino_completo = os.path.join (destino, os.path.basename (origem))
                                        shutil.move (origem, destino_completo)
                                        print(f'Salvo na pasta: {pasta_final}')


                                    page1.wait_for_timeout(3000)
                                    # ---------------------
                                    context.close()
                                    browser.close()

            def Sicredi(playwright: Playwright) -> None:
                print()
                print("Sicredi")
                print()
                
                workbook = openpyxl.load_workbook("C:\python\CONTABILIDADE\Pasta1.xlsx")

                planilha = workbook['Planilha1']
                planilha = workbook.active  
                
                Linha_Parada = 0

                for row in planilha.iter_rows(min_col=5, max_col=5):
                    celula = row[0]
                    if celula.value =='Sicredi':
                        ultima_linha = celula.row

                print(ultima_linha)

                for row in range(1, ultima_linha + 1):
                    
                    print(row)
                #Coluna nome da empresa
                    celula = planilha.cell(row=row, column=1)
                    empresa = celula.value

                    #COLUNA CNPJ/CPF
                    celula = planilha.cell(row=row, column=2)
                    CNPJ_CPF = celula.value

                    #Coluna login
                    celula = planilha.cell(row=row, column=3)
                    Login = celula.value

                    #coluna senha
                    celula = planilha.cell(row=row, column=4)
                    SenhaB = celula.value
                    SenhaB = str(SenhaB)

                    #coluna banco
                    celula = planilha.cell(row=row, column=5)
                    BancoPlanilha = celula.value


                    if empresa == Empresa:
                        if BancoPlanilha == 'Sicredi':
                            print(empresa)
                            CNPJ_CPF = CNPJ_CPF.replace('.','')
                            CNPJ_CPF = CNPJ_CPF.replace('/','')
                            CNPJ_CPF = CNPJ_CPF.replace('-','')

                            browser = playwright.chromium.launch(channel="chrome", headless=headless)
                            context = browser.new_context()
                            page = context.new_page()
                            page.goto("https://www.sicredi.com.br/home/")
                            if MostrarExecução == "1":
                                page.get_by_role("button", name="Permitir Todos").click()
                            
                            page.get_by_role("button", name="5D1A52AE-EC5E-46CD-ACFA-BEDFB093ED60 Acessar minha conta").click()
                            page.get_by_placeholder("CNPJ").click()
                            page.get_by_placeholder("CNPJ").fill(CNPJ_CPF)
                            page.get_by_placeholder("CNPJ").press("Enter")
                            page.get_by_role("button", name="Acessar", exact=True).click()
                            page.get_by_placeholder("Ex.:usuario").click()
                            page.get_by_placeholder("Ex.:usuario").fill(Login)
                            #criar lista para indentificação dos click da senha
                            senha = list(SenhaB)
                            Button0 = page.get_by_role("button", name="Acessar", exact=True).text_content
                            p = list(page.get_by_title("Inserir o digito").all_text_contents()) #lista de botoes
                            
                            
                            #Executa os clicks de acordo com a senha
                            for n in senha:
                                for s in p:
                                    if n in s:
                                        page.wait_for_timeout(1000)
                                        page.get_by_text(s).click()
                                        #print(s)

                            #Finaliza os acessos ate a Cosulta
                            page.get_by_text("Acessar").click()
                            page.get_by_role("link", name="Extrato").click()
                            page.locator("#dataInicialExtrato").click()
                            #Data inicial e final de acordo com o periodo da variavel
                            l = 0
                            for l in lista_de_mes:
                                if l < 10:
                                    l = "0" + str(l)
                                else:
                                    l = int(l)        
                                print(l)
                                AnoDoMes = datetime.now().year
                                limitador = [] 
                                if MesLimite < 10:
                                    limitador = "0" + str(MesLimite) + "/" + str(Ano)
                                else:
                                    limitador = str(MesLimite) + "/" + str(Ano)
                                #print(limitador)
                                res = calendar.monthrange(int(Ano), int(l))
                                res[0] 
                                res[1]
                                ultimoDia = res[1]
                                DataInicial = "01" + "/" + str(l) + "/" + str(Ano)
                                DataFinal = str(ultimoDia) + "/" + str(l) + "/" + str(Ano)
                                MesdeSelecao = str(l) + "/" + str(Ano)
                                #verifivação se a data esta antes do peiodo mensal ou não
                                if limitador == MesdeSelecao:
                                    #EXTRATO POR DATA 
                                        page.get_by_role("button", name="Movimentações Recentes").click()
                                        page.locator("#dataInicialExtrato").fill(DataInicial)
                                        page.locator("#dataInicialExtrato").press("Tab")
                                        page.locator("#dataFinalExtrato").fill(DataFinal)
                                        #page.locator("#dataFinalExtrato").press("Enter")
                                        page.get_by_role("button", name=" Pesquisar").click()
                                        if int(l) < 10:
                                            l = l.replace("0", "")
                                        page.wait_for_timeout(5000)

                                        #ofx
                                        with page.expect_download() as download_info:
                                                page.get_by_role("button", name=" Gerar OFX").click()
                                        download = download_info.value
                                        diretorio = os.getcwd()
                                        DataL = str(l) + "/" + str(Ano)
                                        pasta_final = os.path.join(diretorio, f"Área de Trabalho/extrato.pdf")
                                        pasta_final = os.path.join(diretorio, f"download/sicredi/{empresa.replace(' ', '_')}/OFX/{DataL.replace('/','')}.ofx")
                                        pasta_final = pasta_final.replace("\\", '/')
                                        download.save_as(pasta_final)
                                        origem = pasta_final
                                        destino = f"{Salvar.replace(';',':')} /sicredi/{empresa.replace(' ', '_')}/OFX/"
                                        destino = destino.replace(' ', '')
                                        caminho = Path (destino)
                                        existe = caminho.exists () # retorna True ou False
                                        print(existe)
                                        if existe == False:
                                            os.makedirs (destino)
                                        destino_completo = os.path.join (destino, os.path.basename (origem))
                                        shutil.move (origem, destino_completo)
                                        print(f'Salvo na pasta: {destino_completo}')
                                        page.wait_for_timeout(2000)

                                        #PDF
                                        with page.expect_download() as download1_info:
                                            page.get_by_role("button", name=" Gerar PDF").click()
                                        download = download1_info.value
                                        diretorio = os.getcwd()
                                        DataL = str(l) + "/" + str(Ano)
                                        pasta_final = os.path.join(diretorio, f"Área de Trabalho/extrato.pdf")
                                        pasta_final = os.path.join(diretorio, f"download/sicredi/{empresa.replace(' ', '_')}/pdf/{DataL.replace('/','')}.pdf")
                                        download.save_as(pasta_final)
                                        origem = pasta_final
                                        destino = f"{Salvar.replace(';',':')}/sicredi/{empresa.replace(' ', '_')}/PDF/"
                                        destino = destino.replace(' ', '')
                                        caminho = Path (destino)
                                        existe = caminho.exists () # retorna True ou False
                                        print(existe)
                                        if existe == False:
                                            os.makedirs (destino)     
                                        destino_completo = os.path.join (destino, os.path.basename (origem))
                                        shutil.move (origem, destino_completo)
                                        print(f'Salvo na pasta: {destino_completo}')

                                        #excel
                                        with page.expect_download() as download_info:
                                            page.get_by_role("button", name=" Gerar Planilha").click()
                                        download = download_info.value
                                        diretorio = os.getcwd()
                                        DataL = str(l) + "/" + str(Ano)
                                        pasta_final = os.path.join(diretorio, f"Área de Trabalho/extrato.pdf")
                                        pasta_final = os.path.join(diretorio, f"download/sicredi/{empresa.replace(' ', '_')}/EXCEL/{DataL.replace('/','')}.xls")
                                        pasta_final = pasta_final.replace("\\", '/')
                                        download.save_as(pasta_final)
                                        origem = pasta_final
                                        destino = f"{Salvar.replace(';',':')}/sicredi/{empresa.replace(' ', '_')}/EXCEL/"
                                        destino = destino.replace(' ', '')
                                        caminho = Path (destino)
                                        existe = caminho.exists () # retorna True ou False
                                        print(existe)
                                        if existe == False:
                                            os.makedirs (destino) 
                                        destino_completo = os.path.join (destino, os.path.basename (origem))
                                        shutil.move (origem, destino_completo)
                                        print(f'Salvo na pasta: {destino_completo}')
                                else:
                                        page.get_by_role("button", name="Movimentações Anteriores").click()
                                        page.locator("#comboAnos #opcoes").select_option(Ano)
                                        page.wait_for_timeout(2000)                         
                                        #lista de meses da Selecione o mes
                                        ListaMes = []
                                        options = page.query_selector_all('#comboMeses #opcoes > option')
                                        for option in options:
                                            value = option.get_attribute('value')
                                            if int(value) < 10:
                                                Value = "0" + str(value)
                                                ListaMes.append(value)
                                            else:
                                                ListaMes.append(value)

                                        print(ListaMes)
                                        if int(l) < 10:
                                            l = l.replace("0", "")

                                        l= str(l)
                                        resultado = l in ListaMes                
                                        if resultado == True:    
                                            posicao = ListaMes.index(str(l)) 
                                                                    
                                            page.locator("#comboMeses #opcoes").select_option(str(l))
                                            #page.locator("#comboMeses #opcoes").select_option(str(posicao))
                                            page.wait_for_timeout(3000)
                                            page.get_by_role("button", name=" Pesquisar").click()
                                            page.wait_for_timeout(5000)

                                            #ofx
                                            with page.expect_download() as download_info:
                                                page.get_by_role("button", name=" Gerar OFX").click()
                                            download = download_info.value
                                            diretorio = os.getcwd()
                                            DataL = str(l) + "/" + str(Ano)
                                            pasta_final = os.path.join(diretorio, f"Área de Trabalho\extrato.pdf")
                                            pasta_final = os.path.join(diretorio, f"download/sicredi/{empresa.replace(' ', '_')}/OFx/{DataL.replace('/','')}.ofx")
                                            pasta_final = pasta_final.replace("\\", '/')
                                            download.save_as(str(pasta_final))
                                            origem = pasta_final
                                            destino = f"{Salvar.replace(';',':')}/sicredi/{empresa.replace(' ', '_')}/OFX/"
                                            destino = destino.replace(' ', '')
                                            caminho = Path (destino)
                                            existe = caminho.exists () # retorna True ou False
                                            print(existe)
                                            if existe == False:
                                                os.makedirs (destino)   
                                            destino_completo = os.path.join (destino, os.path.basename (origem))
                                            shutil.move (origem, destino_completo)
                                            print(f'Salvo na pasta: {destino_completo}')
                                            page.wait_for_timeout(2000)

                                            #PDF
                                            with page.expect_download() as download1_info:
                                                page.get_by_role("button", name=" Gerar PDF").click()
                                            download = download1_info.value
                                            diretorio = os.getcwd()
                                            DataL = str(l) + "/" + str(Ano)
                                            pasta_final = os.path.join(diretorio, f"Área de Trabalho/extrato.pdf")
                                            pasta_final = os.path.join(diretorio, f"download/sicredi/{empresa.replace(' ', '_')}/PDF/{DataL.replace('/','')}.pdf")
                                            pasta_final = pasta_final.replace("\\", '/')
                                            print(pasta_final)
                                            download.save_as(pasta_final)
                                            origem = pasta_final
                                            destino = f"{Salvar.replace(';',':')}/sicredi/{empresa.replace(' ', '_')}/PDF/"
                                            destino = destino.replace(' ', '')
                                            caminho = Path (destino)
                                            existe = caminho.exists () # retorna True ou False
                                            print(existe)
                                            if existe == False:
                                                os.makedirs (destino) 
                                            destino_completo = os.path.join (destino, os.path.basename (origem))
                                            shutil.move (origem, destino_completo)
                                            print(f'Salvo na pasta: {destino_completo}')

                                        #excel
                                            with page.expect_download() as download_info:
                                                page.get_by_role("button", name=" Gerar Planilha").click()
                                            download = download_info.value
                                            diretorio = os.getcwd()
                                            DataL = str(l) + "/" + str(Ano)
                                            pasta_final = os.path.join(diretorio, f"Área de Trabalho/extrato.pdf")
                                            pasta_final = os.path.join(diretorio, f"download/sicredi/{empresa.replace(' ', '_')}/EXCEL/{DataL.replace('/','')}.xls")
                                            pasta_final = pasta_final.replace("\\", '/')
                                            download.save_as(pasta_final)
                                            origem = pasta_final
                                            destino = f"{Salvar.replace(';',':')}/sicredi/{empresa.replace(' ', '_')}/EXCEL/"
                                            destino = destino.replace(' ', '')
                                            caminho = Path (destino)
                                            existe = caminho.exists () # retorna True ou False
                                            print(existe)
                                            if existe == False:
                                                os.makedirs (destino) 
                                            destino_completo = os.path.join (destino, os.path.basename (origem))
                                            shutil.move (origem, destino_completo)
                                            print(f'Salvo na pasta: {destino_completo}')

                    #page.wait_for_timeout(1000)
                    # ---------------------
                    #context.close()
                    #browser.close()

                    if row == ultima_linha:
                        print('sestema fechado')
                        Gerados()                 
                    else:
                        print('AINDA NÃO É A LINHA')



            def Itau(playwright: Playwright) -> None:
                print()
                print("ITAU")
                print()
                workbook = openpyxl.load_workbook("C:\python\CONTABILIDADE\Pasta1.xlsx")

                planilha = workbook['Planilha1']
                planilha = workbook.active
                    
                for row in range(2, planilha.max_row + 1):
            
                #Coluna nome da empresa
                    celula = planilha.cell(row=row, column=1)
                    empresa = celula.value

                    #COLUNA CNPJ/CPF
                    celula = planilha.cell(row=row, column=2)
                    CNPJ_CPF = celula.value

                    #Coluna login
                    celula = planilha.cell(row=row, column=3)
                    Login = celula.value
                    Login = str(Login)

                    #coluna senha
                    celula = planilha.cell(row=row, column=4)
                    SenhaB = celula.value
                    SenhaB = str(SenhaB)

                    #coluna banco
                    celula = planilha.cell(row=row, column=5)
                    BancoPlanilha = celula.value


                    if empresa == Empresa:
                        if BancoPlanilha == 'Itau':

                            print(empresa)
                            CNPJ_CPF = CNPJ_CPF.replace('.','')
                            CNPJ_CPF = CNPJ_CPF.replace('/','')
                            CNPJ_CPF = CNPJ_CPF.replace('-','')

                            browser = playwright.chromium.launch(channel="chrome", headless=headless)
                            context = browser.new_context()
                            page = context.new_page()
                            page.goto("https://www.itau.com.br/")
                            page.get_by_role("button", name="ok").click()
                            page.get_by_role("button", name="Mais acessos").click()
                            page.get_by_role("dialog", name="acesse sua conta").locator("div").filter(has_text="acesse sua conta opções de acesso agência e conta código do operador CPF Mais ac").nth(2).click()
                            page.wait_for_timeout(2000)
                            page.get_by_role("combobox", name="opções de acesso").select_option("0: agencia_conta")
                            page.wait_for_timeout(2000)
                            page.get_by_role("combobox", name="opções de acesso").select_option("1: codigo_operador")
                            page.get_by_role("textbox", name="código do operador").click()
                            page.get_by_role("textbox", name="código do operador").fill(Login)
                            page.wait_for_timeout(2000)
                            page.get_by_role("combobox", name="opções de acesso").select_option("0: agencia_conta")
                            page.wait_for_timeout(2000)
                            page.get_by_role("combobox", name="opções de acesso").select_option("1: codigo_operador")
                            page.wait_for_timeout(2000)
                            page.get_by_role("button", name="acessar", exact=True).click()
                            page.get_by_placeholder("senha eletrônica").click()

                            page.wait_for_timeout(3000)
                            listaBotao = list(page.get_by_role("button").all_text_contents())
                            #print(listaBotao)
                            senha = list(SenhaB)
                            lista_de_mes = list(range(int(MesInicial), int(MesFinal)+1))
                            l = 0
                            #Executa os clicks de acordo com a senha
                            for n in senha:
                                    
                                    for s in listaBotao:
                                        if n in s:
                                            page.wait_for_timeout(1000)
                                            page.get_by_text(s).click()
                                            #print(s)
                        
                            page.get_by_role("button", name="acessar").click()
                            page.locator("#rdBasico").check()
                            page.get_by_role("button", name="Continuar").click()
                            page.get_by_role("button", name="ver extrato").click()
                            page.get_by_role("combobox", name="filtrar por período").select_option("mesAtual")
                            page.get_by_role("combobox", name="ordenar por").select_option("asc")
                            page.get_by_role("combobox", name="filtrar por período").select_option("mesCompleto")

                            lista_de_mes = list(range(int(MesInicial), int(MesFinal)+1))
                            l = 0
                            for l in lista_de_mes:
                                print(l)
                                if l < 10:
                                        DataL = "0" + str(l) +"/"+ Ano  + "_"
                                        l = "0" + str(l)
                                else:
                                        DataL = str(l) +"/"+ Ano + "_"
                                #print(DataL)

                                page.get_by_role("button", name="personalizar, mês e ano").click()
                                page.get_by_placeholder("__/____").click()
                                page.get_by_placeholder("__/____").fill(DataL)
                                page.get_by_placeholder("__/____").press("Enter")
                                page.get_by_role("button", name="ok").click()
                                

                                #SALVAR PDF
                                #page.locator("#salvarPdfNovo").get_by_role("button", name="salvar em PDF").click()
                                with page.expect_download() as download_info:
                                    page.locator("#salvarPdfNovo").get_by_role("button", name="salvar em PDF").click()
                                download = download_info.value
                                diretorio = os.getcwd()
                                DataL = str(l) + "/" + str(Ano)
                                pasta_final = os.path.join(diretorio, f"Área de Trabalho/extrato.pdf")
                                pasta_final = os.path.join(diretorio, f"download/ITAU/PDF/{DataL.replace('/','')}.pdf")
                                pasta_final = pasta_final.replace("\\", '/')
                                download.save_as(pasta_final)
                                origem = pasta_final
                                destino = f"{Salvar.replace(';',':')}/ITAU/{empresa.replace(' ', '_')}/PDF/"
                                destino = destino.replace(' ', '')
                                caminho = Path (destino)
                                existe = caminho.exists () # retorna True ou False
                                #print(existe)
                                if existe == False:
                                    os.makedirs (destino) 
                                destino_completo = os.path.join (destino, os.path.basename (origem))
                                shutil.move (origem, destino_completo)
                                print(f'Salvo na pasta: {pasta_final}')

                                #OFX
                                page.get_by_role("button", name="salvar em OFX").click()
                                with page.expect_download() as download_info:
                                    page.get_by_role("button", name="sim, com saldo do dia").click()
                                download = download_info.value
                                diretorio = os.getcwd()
                                DataL = str(l) + "/" + str(Ano)
                                pasta_final = os.path.join(diretorio, f"Área de Trabalho/extrato.pdf")
                                pasta_final = os.path.join(diretorio, f"download/ITAU/OFX/{DataL.replace('/','')}.ofx")
                                pasta_final = pasta_final.replace("\\", '/')
                                download.save_as(pasta_final)
                                origem = pasta_final
                                destino = f"{Salvar.replace(';',':')}/ITAU/{empresa.replace(' ', '_')}/OFX/"
                                destino = destino.replace(' ', '')
                                caminho = Path (destino)
                                existe = caminho.exists() # retorna True ou False
                                #print(existe)
                                if existe == False:
                                    os.makedirs (destino) 
                                destino_completo = os.path.join (destino, os.path.basename (origem))
                                shutil.move (origem, destino_completo)
                                print(f'Salvo na pasta: {pasta_final}')



            #banco = input('Qual banco você quer acessar? ')
            def executa_os_bancos():
            # Chamar a função correspondente ao banco escolhido
                with sync_playwright() as playwright:
                    if Banco == "Todos":
                        Banrisul(playwright)
                        Sicredi(playwright)
                        Itau(playwright)
                    elif Banco == "Itau":  
                        
                        Itau(playwright)  # chama a função itau com o argumento playwright
                    elif Banco == "Banrisul":
                        Banrisul(playwright) # chama a função BANRISUL com o argumento playwright
                    elif Banco == "Sicredi":
                        Sicredi(playwright) # chama a função sicredi com o argumento playwright
                    else:
                        print("Banco inválido") # imprime uma mensagem de erro se o banco não for válido
        
        executa_os_bancos()
        window = tk.Tk()
        window.title("Robson")
        window.geometry("220x100")
        def on_click():
            window.destroy()

        # Defina a cor de fundo para laranja
        window.configure(bg="orange")
        CONCLUIDO_label = tk.Label(window, text="Concluido", bg="orange", font=("Arial", 30, "bold"), fg="white")
        CONCLUIDO_label.grid(row=1, column=0, columnspan=3)

        button = tk.Button(window, text="OK", command=on_click, font=("Arial", 16, "bold"))
        button.grid(row=2, column=1)

            

# Define the window size and title
window = tk.Tk()
window.title("Robson")
window.geometry("450x300")

# Set the background color to orange
window.configure(bg="orange")

NomeUsuario_label = tk.Label(window, text="Nome do usuario:", bg="orange")
NomeUsuario_label.grid(row=0, column=0)

NomeUsuario_input = tk.Entry(window)
NomeUsuario_input.grid(row=0, column=1)

# Create a label for the start month
start_month_label = tk.Label(window, text="Mês inicial:", bg="orange")
start_month_label.grid(row=1, column=0)

# Create a text input box for the start month
start_month_input = tk.Entry(window)
start_month_input.grid(row=1, column=1)

# Create a label for the end month
end_month_label = tk.Label(window, text="Mês final:", bg="orange")
end_month_label.grid(row=2, column=0)

# Create a text input box for the end month
end_month_input = tk.Entry(window)
end_month_input.grid(row=2, column=1)

# Create a label for the year
year_label = tk.Label(window, text="Ano:", bg="orange")
year_label.grid(row=3, column=0)

# Create a text input box for the year
year_input = tk.Entry(window)
year_input.grid(row=3, column=1)

# Create a label for the bank
bank_label = tk.Label(window, text="Banco:", bg="orange")
bank_label.grid(row=4, column=0)

# Create a text input box for the bank
bank_input = ttk.Combobox(window, values=column_k_values)
bank_input.grid(row=4, column=1)

empresas_label = tk.Label(window, text="Empresas:", bg="orange")
empresas_label.grid(row=5, column=0)

empresas_input = ttk.Combobox(window, values=column_a_values)
empresas_input.grid(row=5, column=1)

salvar_label = tk.Label(window, text="Salvar:", bg="orange")
salvar_label.grid(row=6, column=0)

caminho =""
salvar_label = tk.Label(window, text=caminho, bg="orange")
salvar_label.grid(row=6, column=1)

mostrar_label = tk.Label(window, text="Mostrar execução:", bg="orange")
mostrar_label.grid(row=0, column=3, columnspan=2)

#Botoes:
#Create a button to generate the report

generate_report_button = tk.Button(window, text="Gerar relatório", command=salvar)
generate_report_button.grid(row=7, column=1)

Buscar_report_button = tk.Button(window, text="Buscar Arquivos", command = Execucao_Robson)
#generate_report_button.pack()
Buscar_report_button.grid(row=7, column=3)

salvar_report_button = tk.Button(window, text="Buscar pasta", command=buscar_caminho)
salvar_report_button.grid(row=6, column=2, columnspan=2)

AbrirExcel_report_button = tk.Button(window, text="Abrir Excel", command=Abrir_Excel)
AbrirExcel_report_button.grid(row=5, column=2, columnspan=2)

var = tk.IntVar()
#botão sim
sim = tk.Radiobutton(window, text="Sim", variable=var, value=1, bg="orange")
sim.grid(row=1, column=3)

#botão não
nao = tk.Radiobutton(window, text="Não", variable=var, value=0, bg="orange")
nao.grid(row=2, column=3)


#logo da empresa
Expande_label = tk.Label(window, text="Expande", bg="orange", font=("Arial", 70, "bold"), fg="white")
Expande_label.grid(row=8, column=0, columnspan=4)

# Start the main loop
window.mainloop()
