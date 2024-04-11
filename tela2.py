#https://www.youtube.com/watch?v=gjaiAYMhi3U&t=76s
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook
import os

# Load the Excel workbook.
workbook = load_workbook('C:\python\CONTABILIDADE\Pasta1.xlsx')  #Tem que estar na mesma pasta


# Get the "Planilha1" worksheet.
sheet = workbook.get_sheet_by_name('Planilha1')

# Get the values from column A.
column_a_values = [cell.value for cell in sheet['A']]
column_k_values = [cell.value for cell in sheet['L']]

# Inicializar a variável caminho com uma string vazia


def buscar_caminho():
    """Busca o caminho da pasta para salvar um arquivo em Python no Tkinter.

    Returns:
    O caminho da pasta selecionada pelo usuário.
    """
    root = tk.Tk()
    root.withdraw()
    
    diretorio = filedialog.askdirectory(title="Selecione a pasta para salvar o arquivo")
    
    global caminho
    caminho = diretorio  
    print(caminho)

def limpar():
    # Limpar o widget NomeUsuario_input
    NomeUsuario_input.delete(0, tk.END)

    # Limpar o widget start_month_input
    start_month_input.delete(0, tk.END)

    # Limpar o widget end_month_input
    end_month_input.delete(0, tk.END)

    # Limpar o widget year_input
    year_input.delete(0, tk.END)

    # Limpar o widget bank_input
    bank_input.delete(0, tk.END)

    # Limpar o widget empresas_input
    empresas_input.delete(0, tk.END)


def salvar():

    # Pega o input do usuário
    NomeUsuario = NomeUsuario_input.get()
    mes_inicial = start_month_input.get()
    mes_final = end_month_input.get()
    ano = year_input.get()
    banco = bank_input.get()
    empresa = empresas_input.get()
    salvar = caminho
    MostrarExecução = var.get()
    # Cria um arquivo de texto
    with open("S:\Contabilidade\Renan\TESTE\input.txt", "w") as f:
             
        # Escreve o input do usuário no arquivo
        f.write("Nome do usuario: {}\n".format(NomeUsuario))
        f.write("Mês inicial: {}\n".format(mes_inicial))
        f.write("Mês final: {}\n".format(mes_final))
        f.write("Ano: {}\n".format(ano))
        f.write("Banco: {}\n".format(banco))
        f.write("Empresa: {}\n".format(empresa))
        f.write("Salvar: {}\n".format(salvar.replace(":",";")))
        f.write("Mostrar Execução: {}\n".format(MostrarExecução))
        
    window = tk.Tk()
    window.title("Robson")
    window.geometry("220x100")
    def on_click():
        window.destroy()
        
##########################################################################
#Le comandos para o robo
def lerComandos():
     #Ler o txt com os conmandos para execução do robo
        with open("C:\python\CONTABILIDADE\input.txt", "r") as f:
        # Obter uma lista de strings com cada linha do arquivo
            linhas = f.read().splitlines()
        # Criar um dicionário vazio para armazenar as variáveis
        variaveis = {}
        # Percorrer a lista de linhas
        for linha in linhas:
            # Separar o nome e o valor da variável usando o caractere ":"
            nome, valor = linha.split(":")
            # Remover os espaços em branco do nome e do valor
            nome = nome.strip()
            nome = nome.replace("ê", "e")
            nome = nome.replace(" ", "_")
            valor = valor.strip()
            # Converter o valor da variável em seu tipo correspondente usando a função eval
            #valor = eval(valor)
            # Adicionar a variável ao dicionário usando o nome como chave e o valor como valor
            variaveis[nome] = valor
        # Imprimir o dicionário de variáveis na tela
        #print(variaveis)
        MesInicial = variaveis["Mes_inicial"]
        MesFinal = variaveis["Mes_final"]
        Ano = variaveis["Ano"]
        Banco = variaveis["Banco"]
        Empresa = variaveis["Empresa"]
        Salvar = variaveis["Salvar"]
        Salvar = Salvar.replace(";",":")
        MostrarExecução = variaveis["Mostrar_Execução"]

        if MostrarExecução == "1":
            headless = False
        else:
            headless = True

        lista_de_mes = list(range(int(MesInicial), int(MesFinal)+1))
        l = 0

        MesLimite = datetime.now().month - 1 #verifica o mes
        #Acresenta um 0 na frente de numeros menores que 10
        if MesLimite < 10:
            MesLimitel = "0" + str(MesLimite)
        else:
            MesLimite = int(MesLimite)

#leitura da tabela excel
def tabelaexcel():

    empresa = []
    CNPJ_CPF = []
    Login = []
    Senha = []
    BancoPlanilha = []
    Obs = []

    for row in range(2, planilha.max_row + 1):
        #coluna nome da empresa
        celula = planilha.cell(row=row, column=1)
        empresa = celula.value

        #coluna CNPJ_CPF
        celula = planilha.cell(row=row, column=2)
        CNPJ_CPF = celula.value

        #coluna login
        celula = planilha.cell(row=row, column=3)
        Login = celula.value

        #coluna senha
        celula = planilha.cell(row=row, column=4)
        SenhaB = celula.value
        SenhaB = str(SenhaB)

        #coluna banco
        celula = planilha.cell(row=row, column=5)
        BancoPlanilha = celula.value

        #coluna Obs
        celula = planilha.cell(row=row, column=6)
        Obs = celula.value

return empresa, CNPJ_CPF, Login, SenhaB, BancoPlanilha,  





#Funções de navegação dos bancos
#















#########################################################################    
#Função bancos
def Bancos():    
    #Banrisul
    
# Set the background color to orange
    window.configure(bg="orange")
    CONCLUIDO_label = tk.Label(window, text="Concluido", bg="orange", font=("Arial", 30, "bold"), fg="white")
    CONCLUIDO_label.grid(row=1, column=0, columnspan=3)

    button = tk.Button(window, text="OK", command=on_click, font=("Arial", 16, "bold"))
    button.grid(row=2, column=1)



# Define the window size and title
window = tk.Tk()
window.title("Robson")
window.geometry("450x300")

# Set the background color to orange
window.configure(bg="orange")

NomeUsuario_label = tk.Label(window, text="Nome do usuario:", bg="orange")
NomeUsuario_label.grid(row=0, column=0)

NomeUsuario_input = tk.Entry(window)
NomeUsuario_input.grid(row=0, column=1)

# Create a label for the start month
start_month_label = tk.Label(window, text="Mês inicial:", bg="orange")
start_month_label.grid(row=1, column=0)

# Create a text input box for the start month
start_month_input = tk.Entry(window)
start_month_input.grid(row=1, column=1)

# Create a label for the end month
end_month_label = tk.Label(window, text="Mês final:", bg="orange")
end_month_label.grid(row=2, column=0)

# Create a text input box for the end month
end_month_input = tk.Entry(window)
end_month_input.grid(row=2, column=1)

# Create a label for the year
year_label = tk.Label(window, text="Ano:", bg="orange")
year_label.grid(row=3, column=0)

# Create a text input box for the year
year_input = tk.Entry(window)
year_input.grid(row=3, column=1)

# Create a label for the bank
bank_label = tk.Label(window, text="Banco:", bg="orange")
bank_label.grid(row=4, column=0)

# Create a text input box for the bank
bank_input = ttk.Combobox(window, values=column_k_values)
bank_input.grid(row=4, column=1)

empresas_label = tk.Label(window, text="Empresas:", bg="orange")
empresas_label.grid(row=5, column=0)

empresas_input = ttk.Combobox(window, values=column_a_values)
empresas_input.grid(row=5, column=1)

salvar_label = tk.Label(window, text="Salvar:", bg="orange")
salvar_label.grid(row=6, column=0)

caminho =""
salvar_label = tk.Label(window, text=caminho, bg="orange")
salvar_label.grid(row=6, column=1)

mostrar_label = tk.Label(window, text="Mostrar execução:", bg="orange")
mostrar_label.grid(row=0, column=3, columnspan=2)

#Botoes:
#Create a button to generate the report
generate_report_button = tk.Button(window, text="Gerar relatório", command=salvar)
generate_report_button.grid(row=7, column=1)

salvar_report_button = tk.Button(window, text="Buscar pasta", command=buscar_caminho)
salvar_report_button.grid(row=6, column=2, columnspan=2)

var = tk.IntVar()
#botão sim
sim = tk.Radiobutton(window, text="Sim", variable=var, value=1, bg="orange")
sim.grid(row=1, column=3)

#botão não
nao = tk.Radiobutton(window, text="Não", variable=var, value=0, bg="orange")
nao.grid(row=2, column=3)


#logo da empresa
Expande_label = tk.Label(window, text="Expande", bg="orange", font=("Arial", 70, "bold"), fg="white")
Expande_label.grid(row=8, column=0, columnspan=4)

# Start the main loop
window.mainloop()

